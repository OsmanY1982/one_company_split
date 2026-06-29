"""
导出服务
支持多种格式导出：Excel、CSV、PDF、JSON
"""

import json
import csv
import sqlite3
import os
from typing import Dict, List, Optional
from datetime import datetime


class ExportService:
    """导出服务"""

    def __init__(self, db_path: str = "data/app.db", export_dir: str = "exports"):
        self.db_path = db_path
        self.export_dir = export_dir
        os.makedirs(self.export_dir, exist_ok=True)

    def export_to_csv(self,
                      table_name: str,
                      output_path: Optional[str] = None,
                      fields: Optional[List[str]] = None,
                      where: Optional[str] = None,
                      params: Optional[List] = None) -> Dict:
        """导出为CSV"""
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(self.export_dir, f"{table_name}_{timestamp}.csv")

        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row

                query = f"SELECT * FROM {table_name}"
                if where:
                    query += f" WHERE {where}"

                cursor = conn.execute(query, params or [])
                rows = [dict(row) for row in cursor.fetchall()]

            if not rows:
                return {"success": False, "message": "没有数据可导出"}

            # 使用指定的字段或全部字段
            fieldnames = fields or list(rows[0].keys())

            with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
                writer.writeheader()
                writer.writerows(rows)

            return {
                "success": True,
                "file_path": output_path,
                "row_count": len(rows),
                "format": "csv",
            }

        except Exception as e:
            return {"success": False, "message": f"导出失败: {e}"}

    def export_to_json(self,
                       table_name: str,
                       output_path: Optional[str] = None,
                       where: Optional[str] = None,
                       params: Optional[List] = None) -> Dict:
        """导出为JSON"""
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(self.export_dir, f"{table_name}_{timestamp}.json")

        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row

                query = f"SELECT * FROM {table_name}"
                if where:
                    query += f" WHERE {where}"

                cursor = conn.execute(query, params or [])
                rows = [dict(row) for row in cursor.fetchall()]

            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(rows, f, ensure_ascii=False, indent=2)

            return {
                "success": True,
                "file_path": output_path,
                "row_count": len(rows),
                "format": "json",
            }

        except Exception as e:
            return {"success": False, "message": f"导出失败: {e}"}

    def export_to_excel(self,
                        table_name: str,
                        output_path: Optional[str] = None,
                        where: Optional[str] = None,
                        params: Optional[List] = None) -> Dict:
        """导出为Excel"""
        try:
            import openpyxl
        except ImportError:
            return {"success": False, "message": "请安装 openpyxl 库"}

        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(self.export_dir, f"{table_name}_{timestamp}.xlsx")

        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row

                query = f"SELECT * FROM {table_name}"
                if where:
                    query += f" WHERE {where}"

                cursor = conn.execute(query, params or [])
                rows = [dict(row) for row in cursor.fetchall()]

            if not rows:
                return {"success": False, "message": "没有数据可导出"}

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = table_name

            # 写入表头
            headers = list(rows[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # 写入数据
            for row_idx, row in enumerate(rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(header))

            wb.save(output_path)

            return {
                "success": True,
                "file_path": output_path,
                "row_count": len(rows),
                "format": "xlsx",
            }

        except Exception as e:
            return {"success": False, "message": f"导出失败: {e}"}

    def export_batch(self,
                     tables: List[str],
                     format: str = "csv") -> Dict:
        """批量导出"""
        results = {}
        for table in tables:
            if format == "csv":
                results[table] = self.export_to_csv(table)
            elif format == "json":
                results[table] = self.export_to_json(table)
            elif format == "xlsx":
                results[table] = self.export_to_excel(table)

        return {"success": True, "results": results}

    def get_export_history(self) -> List[Dict]:
        """获取导出历史"""
        history = []
        if os.path.exists(self.export_dir):
            for file in os.listdir(self.export_dir):
                file_path = os.path.join(self.export_dir, file)
                if os.path.isfile(file_path):
                    stat = os.stat(file_path)
                    history.append({
                        "file_name": file,
                        "file_path": file_path,
                        "size": stat.st_size,
                        "created_at": datetime.fromtimestamp(stat.st_ctime).isoformat(),
                    })

        return sorted(history, key=lambda x: x["created_at"], reverse=True)

