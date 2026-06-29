# -*- coding: utf-8 -*-
"""
通用 Excel 导出工具
"""

def export_to_excel(parent, headers, rows, default_filename="导出数据"):
    """
    通用 Excel 导出
    :param parent: 父窗口
    :param headers: 列标题列表
    :param rows: 数据行列表（每行是一个列表/元组）
    :param default_filename: 默认文件名
    """
    from PyQt5.QtWidgets import QFileDialog, QMessageBox
    
    path, _ = QFileDialog.getSaveFileName(
        parent, "导出 Excel", default_filename, "Excel (*.xlsx);;CSV (*.csv)"
    )
    if not path:
        return False
    
    try:
        if path.endswith('.csv'):
            import csv
            with open(path, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                for row in rows:
                    writer.writerow([str(v) if v is not None else '' for v in row])
        else:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            except ImportError:
                QMessageBox.warning(parent, "提示", "请先安装 openpyxl：\npip install openpyxl")
                return False
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # 标题行样式
            header_fill = PatternFill("solid", fgColor="2B6CB0")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_align = Alignment(horizontal="center", vertical="center")
            thin = Side(style='thin', color='D1D5DB')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            # 写标题
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = border
            ws.row_dimensions[1].height = 28
            
            # 写数据
            alt_fill = PatternFill("solid", fgColor="EBF8FF")
            for row_idx, row in enumerate(rows, 2):
                for col_idx, val in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx,
                                   value=str(val) if val is not None else '')
                    cell.alignment = Alignment(vertical="center")
                    cell.border = border
                    if row_idx % 2 == 0:
                        cell.fill = alt_fill
                ws.row_dimensions[row_idx].height = 22
            
            # 自动列宽
            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
            
            wb.save(path)
        
        QMessageBox.information(parent, "导出成功", f"已导出 {len(rows)} 条记录至：\n{path}")
        return True
    except Exception as e:
        QMessageBox.critical(parent, "导出失败", f"导出出错：{e}")
        return False
