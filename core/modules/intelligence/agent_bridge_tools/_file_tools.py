"""文件系统工具 Mixin：read_file / write_file / edit_file / list_directory / search_files"""

import os
import fnmatch


class _FileToolsMixin:
    """文件系统工具注册"""

    # ── 1. read_file（支持 PDF/DOCX/XLSX + 纯文本）──
    def _reg_read_file(self):
        def _read_plain_text(path: str, limit: int) -> dict:
            with open(path, "r", encoding="utf-8") as f:
                lines = f.readlines()[:limit]
            return {
                "content": "".join(lines),
                "total_lines": len(lines),
                "truncated": len(lines) >= limit,
                "path": path,
            }

        def _read_pdf(path: str) -> dict:
            import pdfplumber
            with pdfplumber.open(path) as pdf:
                pages_text = []
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        pages_text.append(text)
            content = "\n\n".join(pages_text)
            lines = content.split("\n")
            return {
                "content": content,
                "total_lines": len(lines),
                "truncated": False,
                "path": path,
                "format": "pdf",
                "pages": len(pdf.pages),
            }

        def _read_docx(path: str) -> dict:
            from docx import Document
            doc = Document(path)
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            content = "\n".join(paragraphs)
            lines = content.split("\n")
            return {
                "content": content,
                "total_lines": len(lines),
                "truncated": False,
                "path": path,
                "format": "docx",
                "paragraphs": len(paragraphs),
            }

        def _read_xlsx(path: str, sheet_name: str = None) -> dict:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            sheets_data = {}
            sheet_names = [sheet_name] if sheet_name else wb.sheetnames
            for sn in sheet_names:
                if sn not in wb.sheetnames:
                    continue
                ws = wb[sn]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append("\t".join(str(c) if c is not None else "" for c in row))
                sheets_data[sn] = "\n".join(rows)
            wb.close()
            # 单 sheet 返回纯文本，多 sheet 返回分段
            if sheet_name:
                content = sheets_data.get(sheet_name, "")
            elif len(sheets_data) == 1:
                content = list(sheets_data.values())[0]
            else:
                parts = []
                for sn, text in sheets_data.items():
                    parts.append(f"=== Sheet: {sn} ===\n{text}")
                content = "\n\n".join(parts)
            lines = content.split("\n")
            return {
                "content": content,
                "total_lines": len(lines),
                "truncated": False,
                "path": path,
                "format": "xlsx",
                "sheets": wb.sheetnames,
            }

        def handler(path: str, limit: int = 200, sheet_name: str = None) -> dict:
            try:
                if not os.path.exists(path):
                    return {"error": f"文件不存在: {path}"}
                ext = os.path.splitext(path)[1].lower()
                # 路由：扩展名 → 解析器
                if ext == ".pdf":
                    return _read_pdf(path)
                elif ext == ".docx":
                    return _read_docx(path)
                elif ext in (".xlsx", ".xlsm"):
                    return _read_xlsx(path, sheet_name=sheet_name)
                else:
                    return _read_plain_text(path, limit)
            except Exception as e:
                import traceback
                # 解析失败时回退到纯文本读取
                fallback_warning = f"[警告: {ext} 格式解析失败，回退为纯文本读取。错误: {e}]"
                try:
                    result = _read_plain_text(path, limit)
                    result["warning"] = fallback_warning
                    result["content"] = fallback_warning + "\n\n" + result["content"]
                    return result
                except Exception as e2:
                    return {"error": f"读取失败（解析错误: {e}，回退也失败: {e2}）"}

        self.registry.register(
            name="read_file",
            description="读取文件内容。支持 PDF/DOCX/XLSX/XLSM + 纯文本（.txt/.py/.md/.json 等）。PDF 用 pdfplumber 提取文本，DOCX 用 python-docx 提取段落，XLSX 用 openpyxl 按 sheet 读取（可指定 sheet_name）",
            parameters={
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "文件绝对路径"},
                    "limit": {"type": "integer", "description": "纯文本模式最大读取行数，默认200", "default": 200},
                    "sheet_name": {"type": "string", "description": "Excel 指定 sheet 名称（可选，不传则读取全部 sheet）"},
                },
                "required": ["path"],
            },
            category="file",
        )(handler)

    # ── 2. write_file ──
    def _reg_write_file(self):
        def handler(path: str, content: str) -> dict:
            try:
                os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
                with open(path, "w", encoding="utf-8") as f:
                    f.write(content)
                return {"success": True, "path": path, "bytes": len(content)}
            except Exception as e:
                return {"error": str(e)}

        self.registry.register(
            name="write_file",
            description="创建或覆盖写入文件（自动创建目录）",
            parameters={
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "文件绝对路径"},
                    "content": {"type": "string", "description": "要写入的全部内容"},
                },
                "required": ["path", "content"],
            },
            category="file",
        )(handler)

    # ── 3. edit_file（精准行级编辑，含 diff 预览）──
    def _reg_edit_file(self):
        import difflib

        def handler(path: str, old_str: str, new_str: str, replace_all: bool = False) -> dict:
            try:
                if not os.path.exists(path):
                    return {"error": f"文件不存在: {path}"}
                with open(path, "r", encoding="utf-8") as f:
                    content = f.read()
                count = content.count(old_str)
                if count == 0:
                    return {"error": f"未找到匹配文本。请确认 old_str 与文件中内容完全一致（含空格/换行）"}
                if not replace_all and count > 1:
                    return {
                        "error": f"找到 {count} 处匹配，请设置 replace_all=true 或提供更精确的 old_str",
                        "matches": count,
                    }

                # ── 生成 unified diff 预览（仅首次替换的 diff）──
                diff_text = ""
                try:
                    modified = content.replace(old_str, new_str, 1)
                    diff_lines = list(difflib.unified_diff(
                        content.splitlines(keepends=True),
                        modified.splitlines(keepends=True),
                        fromfile=os.path.basename(path),
                        tofile=os.path.basename(path),
                    ))
                    diff_text = "".join(diff_lines)
                except Exception:
                    diff_text = "[diff 生成失败]"

                # ── 执行替换 ──
                new_content = content.replace(old_str, new_str) if replace_all else content.replace(old_str, new_str, 1)
                with open(path, "w", encoding="utf-8") as f:
                    f.write(new_content)

                result = {
                    "success": True,
                    "path": path,
                    "replacements": count if replace_all else 1,
                    "old_bytes": len(content),
                    "new_bytes": len(new_content),
                }
                if diff_text:
                    result["diff"] = diff_text
                return result
            except Exception as e:
                return {"error": str(e)}

        self.registry.register(
            name="edit_file",
            description="精准替换文件中的文本片段（行级编辑）。old_str 必须与文件内容完全一致（含空格/换行）",
            parameters={
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "文件绝对路径"},
                    "old_str": {"type": "string", "description": "要替换的原始文本，必须完全匹配"},
                    "new_str": {"type": "string", "description": "替换后的新文本"},
                    "replace_all": {"type": "boolean", "description": "是否替换所有匹配项", "default": False},
                },
                "required": ["path", "old_str", "new_str"],
            },
            category="file",
        )(handler)

    # ── 4. list_directory ──
    def _reg_list_directory(self):
        def handler(path: str, pattern: str = "*") -> dict:
            try:
                if not os.path.isdir(path):
                    return {"error": f"不是有效目录: {path}"}
                items = []
                for entry in sorted(os.listdir(path)):
                    full = os.path.join(path, entry)
                    is_dir = os.path.isdir(full)
                    items.append({
                        "name": entry,
                        "type": "dir" if is_dir else "file",
                        "size": os.path.getsize(full) if not is_dir else 0,
                    })
                if pattern != "*":
                    items = [i for i in items if fnmatch.fnmatch(i["name"], pattern)]
                return {"path": path, "count": len(items), "items": items[:200]}
            except Exception as e:
                return {"error": str(e)}

        self.registry.register(
            name="list_directory",
            description="列出目录内容。支持 fnmatch 过滤（如 *.py, test_*）",
            parameters={
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "目录绝对路径"},
                    "pattern": {"type": "string", "description": "文件名通配符，默认 *", "default": "*"},
                },
                "required": ["path"],
            },
            category="file",
        )(handler)

    # ── 5. search_files（glob 搜索）──
    def _reg_search_files(self):
        def handler(directory: str, pattern: str, recursive: bool = True) -> dict:
            import glob
            try:
                if recursive:
                    search_pattern = os.path.join(directory, "**", pattern)
                else:
                    search_pattern = os.path.join(directory, pattern)
                results = glob.glob(search_pattern, recursive=recursive)
                return {"pattern": pattern, "count": len(results), "files": results[:100]}
            except Exception as e:
                return {"error": str(e)}

        self.registry.register(
            name="search_files",
            description="按通配符模式搜索文件（如 **/*.py 递归搜索所有 .py 文件）",
            parameters={
                "type": "object",
                "properties": {
                    "directory": {"type": "string", "description": "搜索根目录"},
                    "pattern": {"type": "string", "description": "glob 模式（如 *.py, test_*.py, **/*.json）"},
                    "recursive": {"type": "boolean", "description": "是否递归子目录", "default": True},
                },
                "required": ["directory", "pattern"],
            },
            category="file",
        )(handler)
